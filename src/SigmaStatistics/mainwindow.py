#!/usr/bin/env python

import sys
import csv
from openpyxl import load_workbook
from PyQt6.QtWidgets import QApplication, QMainWindow, QPlainTextEdit
# , QMessageBox, QTextEdit, QPushButton

from SigmaStatistics.mainwindow_layout import Ui_MainWindow
from SigmaStatistics.methods import Methods
from SigmaStatistics.resources_blob import *

# activate this flag when build the app using pyinstaller:
BUILD = False


class PlainTextEditDragNDrop(QPlainTextEdit):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            event.ignore()

    def dropEvent(self, event):
        for url in event.mimeData().urls():
            file_path = url.toLocalFile()
            if file_path.endswith('.csv'):
                self.open_csv(file_path)
            elif file_path.endswith('.xlsx'):
                self.open_xlsx(file_path)
            else:
                self.setPlainText("Only CSV and XLSX files are supported.")

    def open_csv(self, file_path):
        with open(file_path, 'r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            content = '\n'.join(['\t'.join(row) for row in reader])
            self.setPlainText(content)

    def open_xlsx(self, file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        content = '\n'.join(['\t'.join(
            [str(cell.value) if cell.value is not None else '' for cell in row]) for row in sheet.iter_rows()])
        self.setPlainText(content)


class MainWindow(QMainWindow, Methods):

    def __init__(self):
        super().__init__()
        self.load_UI()

    def load_UI(self,):
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        # Bind functions to butts:
        self.ui.runAutoButton.clicked.connect(self.runAuto)
        self.ui.runManualButton.clicked.connect(self.runManual)
        # replace the input_field with Drag-n-Drop-able one:
        self.replace_input_field()

    def replace_input_field(self):
        new_input_field = PlainTextEditDragNDrop(self)
        new_input_field.setLineWrapMode(QPlainTextEdit.LineWrapMode.NoWrap)
        placeholder_text = self.ui.input_field.placeholderText()
        layout = self.centralWidget().layout()
        layout.replaceWidget(self.ui.input_field, new_input_field)
        self.ui.input_field.deleteLater()
        self.ui.input_field = new_input_field
        self.ui.input_field.setPlaceholderText(placeholder_text)


def main():

    if BUILD:
        import pyi_splash
    app = QApplication(sys.argv)
    main_window = MainWindow()
    main_window.show()
    if BUILD:
        pyi_splash.close()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()
